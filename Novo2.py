import os
import re
import PyPDF2
from openpyxl import load_workbook, Workbook
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

class ExtratorPDF:
    def __init__(self):
        # Criar janela principal
        self.janela = tk.Tk()
        self.janela.title("Extrator de PDF C6")
        self.janela.geometry("600x400")
        
        # Variáveis para armazenar caminhos
        self.caminho_pdf = tk.StringVar()
        self.caminho_excel = tk.StringVar()
        self.senha_pdf = tk.StringVar(value="0000")  # Valor padrão

        self.criar_interface()

    def criar_interface(self):
        # Frame principal com padding
        frame = ttk.Frame(self.janela, padding="10")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Seleção do arquivo PDF
        ttk.Label(frame, text="Arquivo PDF:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(frame, textvariable=self.caminho_pdf, width=50).grid(row=0, column=1, padx=5)
        ttk.Button(frame, text="Procurar", command=self.selecionar_pdf).grid(row=0, column=2)

        # Seleção do arquivo Excel
        ttk.Label(frame, text="Arquivo Excel:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(frame, textvariable=self.caminho_excel, width=50).grid(row=1, column=1, padx=5)
        ttk.Button(frame, text="Procurar", command=self.selecionar_excel).grid(row=1, column=2)

        # Senha do PDF
        ttk.Label(frame, text="Senha do PDF:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(frame, textvariable=self.senha_pdf, width=20).grid(row=2, column=1, sticky=tk.W, padx=5)

        # Barra de progresso
        self.progresso = ttk.Progressbar(frame, length=400, mode='determinate')
        self.progresso.grid(row=3, column=0, columnspan=3, pady=20)

        # Área de log
        self.log_text = tk.Text(frame, height=10, width=60)
        self.log_text.grid(row=4, column=0, columnspan=3, pady=5)

        # Botão de processar
        ttk.Button(frame, text="Processar", command=self.processar).grid(row=5, column=0, columnspan=3, pady=10)

    def log(self, mensagem):
        self.log_text.insert(tk.END, mensagem + "\n")
        self.log_text.see(tk.END)
        self.janela.update()

    def selecionar_pdf(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o arquivo PDF",
            filetypes=[("Arquivos PDF", "*.pdf")]
        )
        if arquivo:
            self.caminho_pdf.set(arquivo)
            # Sugere nome do arquivo Excel baseado no PDF
            excel_sugerido = os.path.splitext(arquivo)[0] + ".xlsx"
            self.caminho_excel.set(excel_sugerido)

    def selecionar_excel(self):
        arquivo = filedialog.asksaveasfilename(
            title="Salvar arquivo Excel",
            filetypes=[("Arquivos Excel", "*.xlsx")],
            defaultextension=".xlsx"
        )
        if arquivo:
            self.caminho_excel.set(arquivo)

    def processar_texto(self, texto):
        # Log do texto original para debug
        self.log("\nTexto original completo:")
        self.log("=" * 50)
        self.log(texto)
        self.log("=" * 50)
        
        # Remove caracteres de formatação mantendo quebras de linha
        texto = texto.replace('\f', '')
        
        # Divide em linhas e processa cada uma
        linhas_processadas = []
        linhas_originais = texto.split('\n')
        
        self.log(f"\nTotal de linhas originais: {len(linhas_originais)}")
        
        for i, linha in enumerate(linhas_originais):
            linha = linha.strip()
            if linha and not linha.startswith('SALDO') and 'SALDO DISPONIVEL' not in linha:
                linhas_processadas.append(linha)
                
        self.log(f"Total de linhas após processamento inicial: {len(linhas_processadas)}")
        return '\n'.join(linhas_processadas)

    def extrair_informacao(self, linha): 
        self.log(f"\nTentando extrair da linha: '{linha}'")
        
        # Ignora linhas de cabeçalho e saldo
        if any(palavra in linha.upper() for palavra in ['SALDO', 'DATA DESCRIÇÃO', 'INICIAL']):
            self.log("Linha ignorada - cabeçalho ou saldo")
            return None
        
        # Padrão de extração ajustado para cobrir descrições mais variadas
        pattern = r'(\d{2}/\d{2}/\d{4})\s+([A-Za-zÀ-Ú0-9\s\*\-\,\.]+?)\s+(\d{12})\s+([\d,.]+)\s*([CD])'
        
        try:
            match = re.search(pattern, linha)
            if match:
                data = match.group(1)
                descricao = match.group(2).strip()
                codtrans = match.group(3)
                valor = match.group(4).replace(',', '.')
                caractere = match.group(5)
                
                self.log("Extração bem-sucedida!")
                self.log(f"Dados extraídos: {data} | {descricao} | {codtrans} | {valor} | {caractere}")
                
                return (data, descricao, codtrans, valor, caractere)
            else:
                self.log("Falha na extração - Linha não corresponde ao padrão")
                return None
            
        except Exception as e:
            self.log(f"Erro ao extrair linha: {str(e)}")
            return None

    def processar(self):
        try:
            # Verificações iniciais permanecem iguais...

            self.log("Iniciando processamento...")
            self.progresso['value'] = 0

            # Lista para armazenar todas as linhas e seu status
            todas_linhas = []
            linhas_nao_extraidas = []

            # Processar PDF
            with open(self.caminho_pdf.get(), 'rb') as arquivo_pdf:
                self.log("Lendo arquivo PDF...")
                leitor_pdf = PyPDF2.PdfReader(arquivo_pdf)
                
                if leitor_pdf.is_encrypted:
                    self.log("PDF protegido, tentando desbloquear...")
                    leitor_pdf.decrypt(self.senha_pdf.get())

                informacoes = []
                total_paginas = len(leitor_pdf.pages)
                texto_completo = ""

                # Extrair texto de cada página
                for i, pagina in enumerate(leitor_pdf.pages):
                    self.log(f"\nLendo página {i+1} de {total_paginas}...")
                    texto_pagina = pagina.extract_text()
                    
                    # Log do texto da página para debug
                    self.log(f"\nConteúdo da página {i+1}:")
                    self.log("-" * 30)
                    self.log(texto_pagina)
                    self.log("-" * 30)
                    
                    texto_completo += texto_pagina
                    self.progresso['value'] = ((i + 1) / total_paginas) * 50

                # Processar o texto completo
                texto_processado = self.processar_texto(texto_completo)
                self.log("\nProcessando texto extraído...")

                # Processar linha por linha
                linhas = texto_processado.split('\n')
                total_linhas = len(linhas)
                
                self.log(f"\nTotal de linhas para processar: {total_linhas}")
                
                for i, linha in enumerate(linhas):
                    if linha.strip():  # Ignora linhas vazias
                        informacao = self.extrair_informacao(linha)
                        if informacao:
                            informacoes.append(informacao)
                            todas_linhas.append(("EXTRAÍDA", linha))
                        else:
                            linhas_nao_extraidas.append(linha)
                            todas_linhas.append(("NÃO EXTRAÍDA", linha))
                    
                    self.progresso['value'] = 50 + ((i + 1) / total_linhas) * 50
                    self.janela.update()

            # Salvar relatório detalhado
            arquivo_log = os.path.splitext(self.caminho_pdf.get())[0] + "_relatorio_extracao.txt"
            self.log(f"\nSalvando relatório detalhado em: {arquivo_log}")
            with open(arquivo_log, 'w', encoding='utf-8') as f:
                f.write("RELATÓRIO DE EXTRAÇÃO\n")
                f.write("=" * 80 + "\n\n")
                f.write(f"Total de linhas processadas: {len(todas_linhas)}\n")
                f.write(f"Linhas extraídas com sucesso: {len(informacoes)}\n")
                f.write(f"Linhas não extraídas: {len(linhas_nao_extraidas)}\n\n")
                
                f.write("DETALHAMENTO DAS LINHAS\n")
                f.write("=" * 80 + "\n\n")
                for status, linha in todas_linhas:
                    f.write(f"{status}: {linha}\n")
                
                f.write("\n\nLINHAS NÃO EXTRAÍDAS\n")
                f.write("=" * 80 + "\n\n")
                for linha in linhas_nao_extraidas:
                    f.write(f"{linha}\n")

            # Criar/atualizar Excel
            self.log("\nSalvando dados no Excel...")
            try:
                if os.path.exists(self.caminho_excel.get()):
                    planilha = load_workbook(self.caminho_excel.get())
                else:
                    planilha = Workbook()

                if 'Planilha1' not in planilha.sheetnames:
                    planilha.create_sheet('Planilha1')
                planilha_ativa = planilha['Planilha1']

                # Limpar planilha
                planilha_ativa.delete_rows(1, planilha_ativa.max_row + 1)

                # Adicionar cabeçalhos
                headers = ['Data', 'Descrição', 'Código', 'Valor', 'Tipo']
                for col, header in enumerate(headers, start=1):
                    planilha_ativa.cell(row=1, column=col, value=header)

                # Adicionar dados
                for linha_num, linha in enumerate(informacoes, start=2):
                    for coluna_num, valor in enumerate(linha, start=1):
                        planilha_ativa.cell(row=linha_num, column=coluna_num, value=valor)

                planilha.save(self.caminho_excel.get())
                self.log("\nProcesso concluído com sucesso!")
                
                # Mostrar estatísticas
                self.log(f"\nEstatísticas:")
                self.log(f"Total de linhas processadas: {len(informacoes) + len(linhas_nao_extraidas)}")
                self.log(f"Linhas extraídas com sucesso: {len(informacoes)}")
                self.log(f"Linhas não extraídas: {len(linhas_nao_extraidas)}")
                
                messagebox.showinfo("Sucesso", f"Extrato processado com sucesso!\n\nLinhas extraídas: {len(informacoes)}\nLinhas não extraídas: {len(linhas_nao_extraidas)}")

            except Exception as e:
                self.log(f"\nErro ao salvar Excel: {str(e)}")
                messagebox.showerror("Erro", f"Erro ao salvar Excel: {str(e)}")

        except Exception as e:
            self.log(f"\nErro durante o processamento: {str(e)}")
            messagebox.showerror("Erro", f"Erro durante o processamento: {str(e)}")

    def iniciar(self):
        self.janela.mainloop()

if __name__ == "__main__":
    app = ExtratorPDF()
    app.iniciar()