import os
import re
import PyPDF2
from openpyxl import load_workbook, Workbook
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

class ExtratorPDF:
    def __init__(self):
        # Criar janela principal
        self.janela = tk.Tk()
        self.janela.title("Extrator de PDF C6")
        self.janela.geometry("600x400")
        
        # Variáveis para armazenar caminhos
        self.caminho_pdf = tk.StringVar()
        self.caminho_excel = tk.StringVar()
        self.senha_pdf = tk.StringVar(value="0000")  # Valor padrão

        self.criar_interface()

    def criar_interface(self):
        # Frame principal com padding
        frame = ttk.Frame(self.janela, padding="10")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Seleção do arquivo PDF
        ttk.Label(frame, text="Arquivo PDF:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(frame, textvariable=self.caminho_pdf, width=50).grid(row=0, column=1, padx=5)
        ttk.Button(frame, text="Procurar", command=self.selecionar_pdf).grid(row=0, column=2)

        # Seleção do arquivo Excel
        ttk.Label(frame, text="Arquivo Excel:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(frame, textvariable=self.caminho_excel, width=50).grid(row=1, column=1, padx=5)
        ttk.Button(frame, text="Procurar", command=self.selecionar_excel).grid(row=1, column=2)

        # Senha do PDF
        ttk.Label(frame, text="Senha do PDF:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(frame, textvariable=self.senha_pdf, width=20).grid(row=2, column=1, sticky=tk.W, padx=5)

        # Barra de progresso
        self.progresso = ttk.Progressbar(frame, length=400, mode='determinate')
        self.progresso.grid(row=3, column=0, columnspan=3, pady=20)

        # Área de log
        self.log_text = tk.Text(frame, height=10, width=60)
        self.log_text.grid(row=4, column=0, columnspan=3, pady=5)

        # Botão de processar
        ttk.Button(frame, text="Processar", command=self.processar).grid(row=5, column=0, columnspan=3, pady=10)

    def log(self, mensagem):
        self.log_text.insert(tk.END, mensagem + "\n")
        self.log_text.see(tk.END)
        self.janela.update()

    def selecionar_pdf(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o arquivo PDF",
            filetypes=[("Arquivos PDF", "*.pdf")]
        )
        if arquivo:
            self.caminho_pdf.set(arquivo)
            # Sugere nome do arquivo Excel baseado no PDF
            excel_sugerido = os.path.splitext(arquivo)[0] + ".xlsx"
            self.caminho_excel.set(excel_sugerido)

    def selecionar_excel(self):
        arquivo = filedialog.asksaveasfilename(
            title="Salvar arquivo Excel",
            filetypes=[("Arquivos Excel", "*.xlsx")],
            defaultextension=".xlsx"
        )
        if arquivo:
            self.caminho_excel.set(arquivo)

    def extrair_informacao(self, linha):
        pattern = r'(\d{2}/\d{2}/\d{4})\s+(.*?)\s+(\d{12})\s+([0-9,.]+)\s+([DC])'
        match = re.search(pattern, linha)
        if match:
            data = match.group(1)
            descricao = match.group(2).strip()
            codtrans = match.group(3)
            valor = match.group(4)
            caractere = match.group(5)
            return (data, descricao, codtrans, valor, caractere)
        return None

    def processar(self):
        try:
            # Verificar se os arquivos foram selecionados
            if not self.caminho_pdf.get() or not self.caminho_excel.get():
                messagebox.showerror("Erro", "Selecione os arquivos PDF e Excel primeiro!")
                return

            self.log("Iniciando processamento...")
            self.progresso['value'] = 0

            # Processar PDF
            with open(self.caminho_pdf.get(), 'rb') as arquivo_pdf:
                self.log("Lendo arquivo PDF...")
                leitor_pdf = PyPDF2.PdfReader(arquivo_pdf)
                
                if leitor_pdf.is_encrypted:
                    self.log("PDF protegido, tentando desbloquear...")
                    leitor_pdf.decrypt(self.senha_pdf.get())

                informacoes = []
                total_paginas = len(leitor_pdf.pages)

                for i, pagina in enumerate(leitor_pdf.pages):
                    self.log(f"Processando página {i+1} de {total_paginas}...")
                    texto = pagina.extract_text()
                    linhas = texto.split('\n')
                    
                    for linha in linhas:
                        informacao = self.extrair_informacao(linha)
                        if informacao:
                            informacoes.append(informacao)
                    
                    # Atualizar barra de progresso
                    self.progresso['value'] = ((i + 1) / total_paginas) * 100
                    self.janela.update()

            # Criar/atualizar Excel
            self.log("Salvando dados no Excel...")
            try:
                if os.path.exists(self.caminho_excel.get()):
                    planilha = load_workbook(self.caminho_excel.get())
                else:
                    planilha = Workbook()

                if 'Planilha1' not in planilha.sheetnames:
                    planilha.create_sheet('Planilha1')
                planilha_ativa = planilha['Planilha1']

                # Limpar planilha
                planilha_ativa.delete_rows(1, planilha_ativa.max_row + 1)

                # Adicionar cabeçalhos
                headers = ['Data', 'Descrição', 'Código', 'Valor', 'Tipo']
                for col, header in enumerate(headers, start=1):
                    planilha_ativa.cell(row=1, column=col, value=header)

                # Adicionar dados
                for linha_num, linha in enumerate(informacoes, start=2):
                    for coluna_num, valor in enumerate(linha, start=1):
                        planilha_ativa.cell(row=linha_num, column=coluna_num, value=valor)

                planilha.save(self.caminho_excel.get())
                self.log("Processo concluído com sucesso!")
                messagebox.showinfo("Sucesso", "Extrato processado com sucesso!")

            except Exception as e:
                self.log(f"Erro ao salvar Excel: {str(e)}")
                messagebox.showerror("Erro", f"Erro ao salvar Excel: {str(e)}")

        except Exception as e:
            self.log(f"Erro durante o processamento: {str(e)}")
            messagebox.showerror("Erro", f"Erro durante o processamento: {str(e)}")

    def iniciar(self):
        self.janela.mainloop()

if __name__ == "__main__":
    app = ExtratorPDF()
    app.iniciar()